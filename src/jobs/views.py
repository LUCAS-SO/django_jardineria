from django.conf import settings
from django.shortcuts import render
from django.views.generic import ListView, DetailView
from django.db.models import Sum
from django.db.models.functions import TruncMonth
from django.utils.timezone import now
from .models import Job
import os


# Exportar a Excel / CSV
import csv
from django.http import HttpResponse, JsonResponse
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment
from datetime import datetime


# Exportar a PDF
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from io import BytesIO


# ==============================
# Helpers
# ==============================

def format_minutes(total_minutes):
    hours = total_minutes // 60
    minutes = total_minutes % 60
    return hours, minutes

# Application startup time for health check
APP_STARTED_AT = now()

def health_check(request):
    uptime = (now() - APP_STARTED_AT).total_seconds()

    return JsonResponse({
        "status": "ok",
        "uptime": uptime,
        "cold": uptime < 20  # primeros 20 segundos = cold start
    })


# Diccionario para traducir meses al español
MESES_ES = {
    "January": "Enero",
    "February": "Febrero",
    "March": "Marzo",
    "April": "Abril",
    "May": "Mayo",
    "June": "Junio",
    "July": "Julio",
    "August": "Agosto",
    "September": "Septiembre",
    "October": "Octubre",
    "November": "Noviembre",
    "December": "Diciembre"
}


# ==============================
# VISTAS WEB
# ==============================

def splash(request):
    return render(request, "splash.html")


class JobListView(ListView):
    model = Job
    template_name = 'job_list.html'
    context_object_name = 'jobs'
    paginate_by = 10

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Agrupación mensual
        monthly_totals = (
            Job.objects
            .annotate(month=TruncMonth('date'))
            .values('month')
            .annotate(total_minutes=Sum('duration'))
            .order_by('-month')
        )

        for item in monthly_totals:
            hours, minutes = format_minutes(item["total_minutes"])
            item["hours"] = hours
            item["minutes"] = minutes

        context['monthly_totals'] = monthly_totals
        return context


class JobDetailView(DetailView):
    model = Job
    template_name = 'job_detail.html'
    context_object_name = 'job'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        job = self.object

        hours, minutes = format_minutes(self.object.duration)
        context["duration_hours"] = hours
        context["duration_minutes"] = minutes

        context["has_before"] = job.photos.filter(before_after='before').exists()
        context["has_after"] = job.photos.filter(before_after='after').exists()
        context["has_both"] = (
            context["has_before"] and context["has_after"]
        )

        return context


# ==============================
# EXPORTAR CSV
# ==============================

def export_jobs_csv(request):
    response = HttpResponse(content_type='text/csv')
    filename = f"trabajos_{datetime.now().strftime('%Y%m%d_%H%M')}.csv"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    writer = csv.writer(response)
    writer.writerow(['Fecha', 'Locación', 'Duración (min)', 'Descripción'])

    for job in Job.objects.order_by('-date'):
        writer.writerow([
            job.date,
            job.get_location_display(),
            job.duration,
            job.description,
        ])

    return response


# ==============================
# EXPORTAR XLSX
# ==============================

def export_jobs_xlsx(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Trabajos"

    # ==========================
    # ESTILOS
    # ==========================
    title_font = Font(size=16, bold=True)
    subtitle_font = Font(size=12, italic=True)
    bold_font = Font(bold=True)
    right_align = Alignment(horizontal="right")

    # ==========================
    # ENCABEZADO / PORTADA
    # ==========================
    ws["A1"] = "Mantenimiento de Espacios Verdes"
    ws["A1"].font = title_font

    # Mes y año (según fecha actual)
    now = datetime.now()
    mes_en = now.strftime("%B")
    mes_es = MESES_ES.get(mes_en, mes_en)
    mes_anio = f"{mes_es} {now.strftime('%Y')}"

    ws["A2"] = f'Informe de trabajos realizados – {mes_anio}'
    ws["A2"].font = subtitle_font

    ws["A3"] = "Lucas Soria"
    ws["A3"].font = bold_font

    ws["A4"] = f"Fecha de descarga: {datetime.now().strftime('%d/%m/%Y %H:%M')}"

    # ==========================
    # LOGO
    # ==========================
    logo_path = os.path.join(settings.BASE_DIR, "static/img/logo.png")
    if os.path.exists(logo_path):
        logo = XLImage(logo_path)
        logo.width = 80
        logo.height = 80
        ws.add_image(logo, "E1")

    # Espacio antes de la tabla
    ws.append([])
    ws.append([])

    start_table_row = ws.max_row + 1

    # ==========================
    # CABECERA TABLA
    # ==========================
    headers = ['Fecha', 'Locación', 'Duración (min)', 'Duración (hh:mm)', 'Descripción']
    ws.append(headers)

    for col in range(1, len(headers) + 1):
        ws.cell(row=start_table_row, column=col).font = bold_font

    # ==========================
    # DATOS
    # ==========================
    total_minutos = 0

    def minutos_a_horas(mins):
        horas = mins // 60
        minutos = mins % 60
        return f"{horas}h {minutos:02d}m"

    for job in Job.objects.order_by('-date'):
        ws.append([
            job.date.strftime("%Y-%m-%d"),
            job.get_location_display(),
            job.duration,
            minutos_a_horas(job.duration),
            job.description,
        ])
        total_minutos += job.duration

    for row in range(start_table_row + 1, ws.max_row + 1):
        ws[f"D{row}"].alignment = right_align

    # ==========================
    # TOTAL
    # ==========================
    ws.append([])
    ws.append(["", "TOTAL", total_minutos, minutos_a_horas(total_minutos), ""])

    total_row = ws.max_row

    ws[f"B{total_row}"].font = bold_font
    ws[f"C{total_row}"].font = bold_font
    ws[f"D{total_row}"].font = bold_font

    # Alineación a la derecha
    ws[f"C{total_row}"].alignment = right_align
    ws[f"D{total_row}"].alignment = right_align

    # ==========================
    # AJUSTE DE COLUMNAS
    # ==========================
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 40

    # ==========================
    # RESPUESTA HTTP
    # ==========================
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"trabajos_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response


# ==============================
# EXPORTAR PDF
# ==============================

def export_jobs_pdf(request):
    buffer = BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=36,
        leftMargin=36,
        topMargin=36,
        bottomMargin=36
    )

    styles = getSampleStyleSheet()
    elements = []

    # ==========================
    # COLORES
    # ==========================
    PRIMARY_COLOR = colors.HexColor("#045C7C")

    # ==========================
    # ESTILOS PERSONALIZADOS
    # ==========================
    styles.add(ParagraphStyle(
        name="TitleCustom",
        fontSize=18,
        leading=22,
        textColor=PRIMARY_COLOR,
        spaceAfter=12,
        alignment=1  # center
    ))

    styles.add(ParagraphStyle(
        name="SubtitleCustom",
        fontSize=11,
        italic=True,
        spaceAfter=8,
        alignment=1
    ))

    styles.add(ParagraphStyle(
        name="Meta",
        fontSize=10,
        spaceAfter=6,
        alignment=1
    ))

    styles.add(ParagraphStyle(
        name="SectionTitle",
        fontSize=14,
        textColor=PRIMARY_COLOR,
        spaceBefore=16,
        spaceAfter=8
    ))

    styles.add(ParagraphStyle(
    name="TableDescription",
    fontSize=9,
    leading=12,
    ))

    # ==========================
    # PORTADA
    # ==========================
    elements.append(
        Paragraph("Mantenimiento de Espacios Verdes", styles["TitleCustom"])
    )

    now = datetime.now()
    mes_en = now.strftime("%B")
    mes_es = MESES_ES.get(mes_en, mes_en)

    elements.append(
        Paragraph(
            f"Informe de trabajos realizados – {mes_es} {now.strftime('%Y')}",
            styles["SubtitleCustom"]
        )
    )

    elements.append(Paragraph("Lucas Soria", styles["Meta"]))
    elements.append(
        Paragraph(
            f"Fecha de descarga: {now.strftime('%d/%m/%Y %H:%M')}",
            styles["Meta"]
        )
    )

    elements.append(Spacer(1, 20))

    # ==========================
    # TABLA DE TRABAJOS
    # ==========================
    jobs = Job.objects.all().order_by("-date")

    data = [["Fecha", "Descripción", "Duración"]]

    total_minutes_all = 0

    for job in jobs:
        dur = int(job.duration) if job.duration else 0
        total_minutes_all += dur

        desc_text = job.description or "N/A"

        data.append([
            job.date.strftime("%d/%m/%Y"),
            Paragraph(desc_text, styles["TableDescription"]),
            f"{dur // 60}h {dur % 60:02d}m"
        ])

    table = Table(data, colWidths=[80, 300, 80])
    table.setStyle(TableStyle([
        # Header
        ('BACKGROUND', (0, 0), (-1, 0), PRIMARY_COLOR),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN', (2, 0), (2, -1), 'RIGHT'),

        # Body
        ('BACKGROUND', (0, 1), (-1, -1), colors.whitesmoke),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
    ]))

    elements.append(table)

    # ==========================
    # TOTALES MENSUALES
    # ==========================
    monthly_totals = (
        Job.objects
        .annotate(month=TruncMonth("date"))
        .values("month")
        .annotate(total_minutes=Sum("duration"))
        .order_by("-month")
    )

    if monthly_totals:
        elements.append(Spacer(1, 16))
        elements.append(Paragraph("Totales Mensuales", styles["SectionTitle"]))

        for item in monthly_totals:
            month = item["month"]
            total_minutes = item["total_minutes"] or 0

            h = total_minutes // 60
            m = total_minutes % 60

            mes_en = month.strftime("%B")
            mes_es = MESES_ES.get(mes_en, mes_en)
            month_str = f"{mes_es} {month.strftime('%Y')}"

            elements.append(
                Paragraph(
                    f"<b>{month_str}:</b> {h}h {m:02d}m",
                    styles["Normal"]
                )
            )

    # ==========================
    # TOTAL GENERAL
    # ==========================
    if total_minutes_all > 0:
        h = total_minutes_all // 60
        m = total_minutes_all % 60

        elements.append(Spacer(1, 10))
        elements.append(
            Paragraph(
                f"Total General: {h}h {m:02d}m",
                styles["SectionTitle"]
            )
        )

    # ==========================
    # CONSTRUIR PDF
    # ==========================
    doc.build(elements)

    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = (
        f'attachment; filename="trabajos_{now.strftime("%Y%m%d_%H%M")}.pdf"'
    )

    return response